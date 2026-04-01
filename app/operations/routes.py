"""Rutas del modulo de operacion contable.

Este archivo concentra la parte mas grande del proceso:
- cumplimiento SAT/EFOS,
- entregables con evidencia,
- traspasos,
- CFDI,
- conciliacion,
- presupuestos,
- reportes,
- acuses/firma.
"""

import os
import re
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from flask import current_app, flash, redirect, render_template, request, send_from_directory, session, url_for
from werkzeug.utils import secure_filename

from app.audit import log_action
from app.auth.routes import login_required, roles_required
from app.extensions import db
from app.models import (
    AuthorizedAccount,
    BankMovement,
    BudgetConfig,
    CfdiRecord,
    EvidenceFile,
    Folio,
    FolioDeliverableItem,
    MonthlyDirectionReport,
    Provider,
    ProviderCompliance,
    ReconciliationAlert,
    SignatureAck,
    Transfer,
)
from app.operations import operations_bp
from app.operations.services import (
    add_business_days,
    ensure_deliverables,
    ensure_workflow,
    kpi_status_for_item,
    update_folio_operational_status,
)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

STATUS_LABELS = {
    "no_cargada": "No cargada",
    "vigente": "Vigente",
    "vencida": "Vencida",
}


def _parse_date(value):
    """Convierte `YYYY-MM-DD` a objeto `date`.

    Si viene vacio, regresa `None`.
    """
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def _store_uploaded_file(file_storage, category, folio_id=None, deliverable_id=None):
    """Guarda un archivo fisico y crea su registro en BD."""
    if not file_storage or not file_storage.filename:
        return None

    safe_name = secure_filename(file_storage.filename)
    ext = os.path.splitext(safe_name)[1]
    stored_name = f"{uuid.uuid4().hex}{ext}"
    save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], stored_name)
    file_storage.save(save_path)

    evidence = EvidenceFile(
        folio_id=folio_id,
        deliverable_id=deliverable_id,
        category=category,
        original_name=safe_name,
        stored_name=stored_name,
        mime_type=file_storage.mimetype,
        uploaded_by=session.get("user_id"),
    )
    db.session.add(evidence)
    return evidence


def _refresh_folio_status(folio):
    """Recalcula workflow + estatus del folio segun sus datos actuales."""
    workflow = ensure_workflow(folio)
    items = FolioDeliverableItem.query.filter_by(folio_id=folio.id).all()
    has_transfer = Transfer.query.filter_by(folio_id=folio.id).count() > 0
    has_cfdi = CfdiRecord.query.filter_by(folio_id=folio.id, status="vigente").count() > 0
    has_reconciliation = BankMovement.query.filter_by(folio_id=folio.id, is_reconciled=True).count() > 0
    update_folio_operational_status(
        folio=folio,
        workflow=workflow,
        deliverables=items,
        has_transfer=has_transfer,
        has_cfdi=has_cfdi,
        has_reconciliation=has_reconciliation,
    )
    db.session.add(workflow)
    db.session.add(folio)


@operations_bp.route("/")
@login_required
def index():
    """Home del modulo Operacion (KPIs y accesos rapidos)."""
    open_alerts = ReconciliationAlert.query.filter_by(status="abierta").count()
    high_alerts = ReconciliationAlert.query.filter_by(status="abierta", severity="alta").count()
    pending_cfdi = Folio.query.filter(Folio.status.in_(["pendiente", "en_proceso", "alerta", "critico"])).count()
    reports = MonthlyDirectionReport.query.order_by(MonthlyDirectionReport.created_at.desc()).limit(6).all()
    signatures = SignatureAck.query.order_by(SignatureAck.signed_at.desc()).all()
    return render_template(
        "operations/index.html",
        open_alerts=open_alerts,
        high_alerts=high_alerts,
        pending_cfdi=pending_cfdi,
        reports=reports,
        signatures=signatures,
    )


@operations_bp.route("/evidencias/<int:evidence_id>")
@login_required
def download_evidence(evidence_id):
    """Descarga un archivo de evidencia previamente cargado."""
    evidence = EvidenceFile.query.get_or_404(evidence_id)
    return send_from_directory(
        current_app.config["UPLOAD_FOLDER"],
        evidence.stored_name,
        as_attachment=True,
        download_name=evidence.original_name,
    )


@operations_bp.route("/cumplimiento", methods=["GET", "POST"])
@login_required
@roles_required("contabilidad", "direccion")
def compliance():
    """Alta/edicion de cumplimiento SAT por proveedor."""
    if request.method == "POST":
        provider_id = int(request.form.get("provider_id", "0"))
        provider = Provider.query.get(provider_id)
        if not provider:
            flash("Proveedor invalido.", "error")
            return redirect(url_for("operations.compliance"))

        item = ProviderCompliance.query.filter_by(provider_id=provider.id).first()
        if not item:
            item = ProviderCompliance(provider_id=provider.id)
            db.session.add(item)

        item.sat_opinion_status = request.form.get("sat_opinion_status", "no_cargada")
        item.sat_valid_until = _parse_date(request.form.get("sat_valid_until"))
        item.efos_flag = bool(request.form.get("efos_flag"))
        item.activity_match = bool(request.form.get("activity_match"))
        item.rfc_matches_account = bool(request.form.get("rfc_matches_account"))
        item.notes = request.form.get("notes", "").strip() or None
        item.updated_at = datetime.now(timezone.utc)
        _store_uploaded_file(request.files.get("sat_file"), "sat_opinion")

        db.session.commit()

        for folio in Folio.query.filter_by(provider_id=provider.id).all():
            _refresh_folio_status(folio)
        db.session.commit()

        log_action(
            user_id=session.get("user_id"),
            action="provider_compliance_updated",
            entity="provider",
            entity_id=provider.id,
            details=f"sat={item.sat_opinion_status};efos={item.efos_flag};activity={item.activity_match};rfc={item.rfc_matches_account}",
        )
        flash("Cumplimiento SAT actualizado.", "success")
        return redirect(url_for("operations.compliance"))

    providers = Provider.query.order_by(Provider.name.asc()).all()
    compliance_map = {item.provider_id: item for item in ProviderCompliance.query.order_by(ProviderCompliance.updated_at.desc()).all()}
    return render_template("operations/compliance.html", providers=providers, compliance_map=compliance_map, status_labels=STATUS_LABELS)


@operations_bp.route("/entregables/<int:folio_id>", methods=["GET", "POST"])
@login_required
@roles_required("ui", "contabilidad", "direccion")
def deliverables(folio_id):
    """Gestion del checklist de entregables por folio."""
    folio = Folio.query.get_or_404(folio_id)
    workflow = ensure_workflow(folio)
    db.session.add(workflow)
    for item in ensure_deliverables(folio):
        db.session.add(item)
    db.session.commit()

    if request.method == "POST":
        deliverable_id = int(request.form.get("deliverable_id", "0"))
        item = FolioDeliverableItem.query.filter_by(id=deliverable_id, folio_id=folio.id).first()
        if not item:
            flash("Entregable invalido.", "error")
            return redirect(url_for("operations.deliverables", folio_id=folio.id))

        evidence = _store_uploaded_file(request.files.get("evidence_file"), "entregable", folio_id=folio.id, deliverable_id=item.id)
        item.uploaded = bool(request.form.get("uploaded")) or evidence is not None
        item.uploaded_at = datetime.now(timezone.utc) if item.uploaded else None
        db.session.add(item)

        _refresh_folio_status(folio)
        db.session.commit()
        log_action(
            user_id=session.get("user_id"),
            action="deliverable_toggled",
            entity="folio",
            entity_id=folio.id,
            details=f"deliverable={item.code};uploaded={item.uploaded};file={evidence.original_name if evidence else 'none'}",
        )
        flash("Entregable actualizado.", "success")
        return redirect(url_for("operations.deliverables", folio_id=folio.id))

    items = FolioDeliverableItem.query.filter_by(folio_id=folio.id).order_by(FolioDeliverableItem.owner_area.asc()).all()
    item_ids = [i.id for i in items]
    evidence_map = {}
    if item_ids:
        for ev in EvidenceFile.query.filter(EvidenceFile.deliverable_id.in_(item_ids)).order_by(EvidenceFile.uploaded_at.desc()).all():
            evidence_map.setdefault(ev.deliverable_id, []).append(ev)
    overdue_count = sum(1 for i in items if kpi_status_for_item(i) == "vencido")

    return render_template(
        "operations/deliverables.html",
        folio=folio,
        items=items,
        workflow=workflow,
        evidence_map=evidence_map,
        kpi_status_for_item=kpi_status_for_item,
        overdue_count=overdue_count,
    )


@operations_bp.route("/traspasos", methods=["GET", "POST"])
@login_required
@roles_required("tesoreria", "contabilidad", "direccion")
def transfers():
    """Registro de traspasos SPEI con reglas de bloqueo."""
    folios = Folio.query.order_by(Folio.created_at.desc()).all()
    accounts = AuthorizedAccount.query.filter_by(is_active=True).all()
    if request.method == "POST":
        folio_id = int(request.form.get("folio_id", "0"))
        account_id = int(request.form.get("account_id", "0"))
        folio = Folio.query.get(folio_id)
        account = AuthorizedAccount.query.get(account_id)
        if not folio or not account:
            flash("Folio o cuenta invalida.", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)

        workflow = ensure_workflow(folio)
        if not workflow.step_2_contab_sat:
            flash("Flujo bloqueado: primero Contabilidad valida SAT/EFOS (Paso 2).", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)
        if account.provider_id != folio.provider_id:
            flash("La cuenta no corresponde al proveedor del folio.", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)

        compliance = ProviderCompliance.query.filter_by(provider_id=folio.provider_id).first()
        if not compliance or compliance.sat_opinion_status != "vigente" or compliance.efos_flag or not compliance.activity_match or not compliance.rfc_matches_account:
            flash("Bloqueo fiscal: SAT/EFOS/actividad/RFC no validado.", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)
        if compliance.sat_valid_until and compliance.sat_valid_until < date.today():
            flash("Opinion SAT vencida. Pago bloqueado.", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)

        try:
            amount = Decimal(request.form.get("amount", "0"))
        except InvalidOperation:
            flash("Monto invalido.", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)

        approved_by_direction = bool(request.form.get("approved_by_direction"))
        if amount > folio.budget_amount and not approved_by_direction:
            flash("Monto excede presupuesto SINGA. Requiere autorizacion de Direccion.", "error")
            return render_template("operations/transfers.html", folios=folios, accounts=accounts)

        transfer = Transfer(
            folio_id=folio.id,
            account_id=account.id,
            spei_reference=request.form.get("spei_reference", "").strip(),
            origin_bank=request.form.get("origin_bank", "").strip(),
            amount=amount,
            transfer_date=_parse_date(request.form.get("transfer_date")) or date.today(),
            approved_by_direction=approved_by_direction,
        )
        db.session.add(transfer)
        if amount > folio.budget_amount:
            db.session.add(ReconciliationAlert(folio_id=folio.id, alert_type="excede_presupuesto", severity="alta", notes="Traspaso mayor al presupuesto autorizado."))
        _refresh_folio_status(folio)
        db.session.commit()

        log_action(user_id=session.get("user_id"), action="transfer_registered", entity="transfer", entity_id=transfer.id, details=f"folio={folio.singa_number};amount={amount}")
        flash("Traspaso registrado.", "success")
        return redirect(url_for("operations.transfers"))

    recent = Transfer.query.order_by(Transfer.created_at.desc()).limit(20).all()
    return render_template("operations/transfers.html", folios=folios, accounts=accounts, recent=recent)


@operations_bp.route("/cfdi", methods=["GET", "POST"])
@login_required
@roles_required("contabilidad", "direccion")
def cfdi():
    """Vinculacion de CFDI (metadatos y archivos XML/PDF)."""
    folios = Folio.query.order_by(Folio.created_at.desc()).all()
    if request.method == "POST":
        folio_id = int(request.form.get("folio_id", "0"))
        folio = Folio.query.get(folio_id)
        if not folio:
            flash("Folio invalido.", "error")
            return render_template("operations/cfdi.html", folios=folios)
        if not ensure_workflow(folio).step_4_tesoreria_transfer:
            flash("Flujo bloqueado: primero Tesoreria registra traspaso (Paso 4).", "error")
            return render_template("operations/cfdi.html", folios=folios)

        try:
            amount = Decimal(request.form.get("amount", "0"))
        except InvalidOperation:
            flash("Monto CFDI invalido.", "error")
            return render_template("operations/cfdi.html", folios=folios)

        record = CfdiRecord(
            folio_id=folio.id,
            uuid=request.form.get("uuid", "").strip(),
            xml_ref=request.form.get("xml_ref", "").strip(),
            pdf_ref=request.form.get("pdf_ref", "").strip(),
            amount=amount,
            issued_at=_parse_date(request.form.get("issued_at")) or date.today(),
            status=request.form.get("status", "vigente"),
            cancel_reason=request.form.get("cancel_reason", "").strip() or None,
        )
        db.session.add(record)
        xml_file = _store_uploaded_file(request.files.get("xml_file"), "cfdi_xml", folio_id=folio.id)
        pdf_file = _store_uploaded_file(request.files.get("pdf_file"), "cfdi_pdf", folio_id=folio.id)
        if xml_file and not record.xml_ref:
            record.xml_ref = xml_file.original_name
        if pdf_file and not record.pdf_ref:
            record.pdf_ref = pdf_file.original_name

        _refresh_folio_status(folio)
        db.session.commit()
        flash("CFDI vinculado al folio.", "success")
        return redirect(url_for("operations.cfdi"))

    records = CfdiRecord.query.order_by(CfdiRecord.created_at.desc()).limit(30).all()
    return render_template("operations/cfdi.html", folios=folios, records=records)


def _parse_bank_pdf(file_storage):
    """Parser base de PDF bancario (extrae lineas con montos)."""
    if PdfReader is None:
        raise RuntimeError("pypdf no esta instalado")
    tmp_name = f"tmp_{uuid.uuid4().hex}.pdf"
    tmp_path = os.path.join(current_app.config["UPLOAD_FOLDER"], tmp_name)
    file_storage.save(tmp_path)
    text = ""
    try:
        reader = PdfReader(tmp_path)
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    parsed = []
    amount_pattern = re.compile(r"([+-]?\d{1,3}(?:,\d{3})*(?:\.\d{2}))")
    for line in [ln.strip() for ln in text.splitlines() if ln.strip()]:
        match = amount_pattern.search(line)
        if not match:
            continue
        try:
            amount = Decimal(match.group(1).replace(",", ""))
        except InvalidOperation:
            continue
        parsed.append(
            {
                "reference": line[:110],
                "amount": abs(amount),
                "movement_type": "deposito" if amount > 0 else "retiro",
            }
        )
    return parsed[:100]


@operations_bp.route("/conciliacion", methods=["GET", "POST"])
@login_required
@roles_required("contabilidad", "direccion")
def reconciliation():
    """Conciliacion bancaria: manual y via carga de PDF."""
    folios = Folio.query.order_by(Folio.created_at.desc()).all()
    if request.method == "POST":
        action = request.form.get("action", "manual")
        if action == "upload_pdf":
            pdf_file = request.files.get("bank_pdf")
            period = request.form.get("period", "").strip()
            bank = request.form.get("bank", "").strip()
            if not pdf_file or not pdf_file.filename:
                flash("Selecciona un PDF de estado de cuenta.", "error")
                return redirect(url_for("operations.reconciliation"))

            evidence = _store_uploaded_file(pdf_file, "bank_statement")
            db.session.flush()
            inserted = 0
            try:
                for item in _parse_bank_pdf(pdf_file):
                    mov = BankMovement(
                        period=period,
                        bank=bank,
                        movement_date=date.today(),
                        movement_type=item["movement_type"],
                        reference=item["reference"],
                        amount=item["amount"],
                        is_reconciled=False,
                    )
                    db.session.add(mov)
                    db.session.add(
                        ReconciliationAlert(
                            movement=mov,
                            alert_type="deposito_sin_folio" if mov.movement_type == "deposito" else "diferencia_cfdi_banco",
                            severity="media",
                            status="abierta",
                            notes="Detectado por carga automatica de PDF bancario.",
                        )
                    )
                    inserted += 1
            except Exception:
                flash("No se pudo extraer texto del PDF, pero se almaceno el archivo.", "error")

            db.session.commit()
            flash(f"PDF cargado ({evidence.original_name}). Movimientos detectados: {inserted}.", "success")
            return redirect(url_for("operations.reconciliation"))

        folio_id_raw = request.form.get("folio_id", "")
        folio = Folio.query.get(int(folio_id_raw)) if folio_id_raw else None
        if folio and not ensure_workflow(folio).step_5_contab_cfdi:
            flash("Flujo bloqueado: primero debe existir CFDI vinculado (Paso 5).", "error")
            return render_template("operations/reconciliation.html", folios=folios)

        try:
            amount = Decimal(request.form.get("amount", "0"))
        except InvalidOperation:
            flash("Monto invalido en movimiento bancario.", "error")
            return render_template("operations/reconciliation.html", folios=folios)

        movement = BankMovement(
            period=request.form.get("period", "").strip(),
            bank=request.form.get("bank", "").strip(),
            movement_date=_parse_date(request.form.get("movement_date")) or date.today(),
            movement_type=request.form.get("movement_type", "retiro"),
            reference=request.form.get("reference", "").strip(),
            amount=amount,
            folio_id=folio.id if folio else None,
            is_reconciled=True,
        )
        db.session.add(movement)
        db.session.flush()

        alerts_created = 0
        account_authorized = bool(request.form.get("account_authorized"))
        if not folio and movement.movement_type == "deposito":
            db.session.add(ReconciliationAlert(movement_id=movement.id, alert_type="deposito_sin_folio", severity="alta", notes="Deposito sin folio SINGA asociado."))
            alerts_created += 1
        if folio and amount > folio.budget_amount:
            db.session.add(ReconciliationAlert(folio_id=folio.id, movement_id=movement.id, alert_type="excede_presupuesto", severity="alta", notes="Movimiento excede presupuesto SINGA."))
            alerts_created += 1
        if not account_authorized:
            db.session.add(ReconciliationAlert(folio_id=folio.id if folio else None, movement_id=movement.id, alert_type="cuenta_no_autorizada", severity="alta", notes="Movimiento detectado a cuenta no autorizada."))
            alerts_created += 1
        if folio and CfdiRecord.query.filter_by(folio_id=folio.id, status="vigente").count() == 0:
            db.session.add(ReconciliationAlert(folio_id=folio.id, movement_id=movement.id, alert_type="diferencia_cfdi_banco", severity="media", notes="Movimiento bancario sin CFDI vigente."))
            alerts_created += 1
        if alerts_created > 0:
            movement.is_reconciled = False
        if folio:
            _refresh_folio_status(folio)
        db.session.commit()
        flash(f"Movimiento conciliado. Alertas generadas: {alerts_created}", "success")
        return redirect(url_for("operations.reconciliation"))

    movements = BankMovement.query.order_by(BankMovement.created_at.desc()).limit(30).all()
    alerts = ReconciliationAlert.query.order_by(ReconciliationAlert.created_at.desc()).limit(30).all()
    return render_template("operations/reconciliation.html", folios=folios, movements=movements, alerts=alerts)


@operations_bp.route("/alertas/<int:alert_id>/resolver", methods=["POST"])
@login_required
@roles_required("contabilidad", "direccion")
def resolve_alert(alert_id):
    """Marca una alerta como resuelta y guarda nota de resolucion."""
    alert = ReconciliationAlert.query.get_or_404(alert_id)
    alert.status = "resuelta"
    alert.resolved_at = datetime.now(timezone.utc)
    resolution = request.form.get("resolution", "").strip()
    if resolution:
        alert.notes = f"{alert.notes or ''} | Resolucion: {resolution}".strip()
    db.session.add(alert)
    db.session.commit()
    flash("Alerta marcada como resuelta.", "success")
    return redirect(url_for("operations.reconciliation"))


@operations_bp.route("/presupuestos", methods=["GET", "POST"])
@login_required
@roles_required("contabilidad", "direccion")
def budgets():
    """Carga de presupuestos SINGA: manual o importacion Excel."""
    folios = Folio.query.order_by(Folio.created_at.desc()).all()
    if request.method == "POST":
        action = request.form.get("action", "single")
        if action == "import_excel":
            if load_workbook is None:
                flash("openpyxl no esta disponible para importar Excel.", "error")
                return redirect(url_for("operations.budgets"))
            excel_file = request.files.get("budget_file")
            if not excel_file or not excel_file.filename:
                flash("Selecciona un archivo Excel.", "error")
                return redirect(url_for("operations.budgets"))

            tmp_name = f"tmp_{uuid.uuid4().hex}.xlsx"
            tmp_path = os.path.join(current_app.config["UPLOAD_FOLDER"], tmp_name)
            excel_file.save(tmp_path)
            loaded = 0
            try:
                wb = load_workbook(tmp_path, data_only=True)
                ws = wb.active
                header = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
                idx = {name: i for i, name in enumerate(header)}
                required = {"folio", "presupuesto autorizado", "periodo"}
                if not required.issubset(set(idx.keys())):
                    flash("El Excel debe incluir columnas: Folio | Presupuesto autorizado | Periodo.", "error")
                    return redirect(url_for("operations.budgets"))
                for row in ws.iter_rows(min_row=2):
                    folio_code = str(row[idx["folio"]].value or "").strip()
                    value = row[idx["presupuesto autorizado"]].value
                    period = str(row[idx["periodo"]].value or "").strip()
                    if not folio_code or value in (None, "") or not period:
                        continue
                    folio = Folio.query.filter_by(singa_number=folio_code).first()
                    if not folio:
                        continue
                    db.session.add(BudgetConfig(folio_id=folio.id, period=period, authorized_budget=Decimal(str(value))))
                    loaded += 1
                db.session.commit()
                flash(f"Carga masiva completada. Registros: {loaded}", "success")
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            return redirect(url_for("operations.budgets"))

        folio_id = int(request.form.get("folio_id", "0"))
        folio = Folio.query.get(folio_id)
        if not folio:
            flash("Folio invalido.", "error")
            return render_template("operations/budgets.html", folios=folios)
        try:
            value = Decimal(request.form.get("authorized_budget", "0"))
        except InvalidOperation:
            flash("Presupuesto invalido.", "error")
            return render_template("operations/budgets.html", folios=folios)
        db.session.add(BudgetConfig(folio_id=folio.id, period=request.form.get("period", "").strip(), authorized_budget=value))
        db.session.commit()
        flash("Presupuesto mensual cargado.", "success")
        return redirect(url_for("operations.budgets"))

    rows = BudgetConfig.query.order_by(BudgetConfig.created_at.desc()).limit(40).all()
    return render_template("operations/budgets.html", folios=folios, rows=rows)


@operations_bp.route("/reportes/generar", methods=["POST"])
@login_required
@roles_required("contabilidad", "direccion")
def generate_report():
    """Genera reporte mensual automatico con KPIs actuales."""
    period = request.form.get("period", "").strip()
    if not period:
        flash("Debes indicar periodo (YYYY-MM).", "error")
        return redirect(url_for("operations.reports"))
    if MonthlyDirectionReport.query.filter_by(period=period).first():
        flash("Ya existe reporte para ese periodo.", "error")
        return redirect(url_for("operations.reports"))

    folios_total = Folio.query.count()
    folios_closed = Folio.query.filter_by(status="cerrado").count()
    open_alerts = ReconciliationAlert.query.filter_by(status="abierta").count()
    high_alerts = ReconciliationAlert.query.filter_by(status="abierta", severity="alta").count()
    sat_vigentes = ProviderCompliance.query.filter_by(sat_opinion_status="vigente", efos_flag=False).count()
    providers_total = Provider.query.count()
    ratio_close = f"{(folios_closed * 100 // folios_total) if folios_total else 0}%"
    sat_ratio = f"{(sat_vigentes * 100 // providers_total) if providers_total else 0}%"

    db.session.add(
        MonthlyDirectionReport(
            period=period,
            executive_summary=f"Folios cerrados: {folios_closed}/{folios_total} ({ratio_close}). Alertas abiertas: {open_alerts} (altas: {high_alerts}).",
            risk_summary=f"SAT vigente/no EFOS: {sat_vigentes}/{providers_total} ({sat_ratio}). Alertas criticas pendientes: {high_alerts}.",
            recommendations="1) Resolver alertas altas <=5 dias habiles. 2) Revalidar opinion SAT previa a pago. 3) Confirmar expedientes completos al dia 27.",
            created_by=session.get("user_id"),
        )
    )
    db.session.commit()
    flash("Reporte mensual autogenerado y almacenado.", "success")
    return redirect(url_for("operations.reports"))


@operations_bp.route("/reportes", methods=["GET", "POST"])
@login_required
@roles_required("contabilidad", "direccion")
def reports():
    """Alta manual/listado de reportes mensuales a Direccion."""
    if request.method == "POST":
        period = request.form.get("period", "").strip()
        if MonthlyDirectionReport.query.filter_by(period=period).first():
            flash("Ya existe reporte para ese periodo.", "error")
            return redirect(url_for("operations.reports"))
        db.session.add(
            MonthlyDirectionReport(
                period=period,
                executive_summary=request.form.get("executive_summary", "").strip(),
                risk_summary=request.form.get("risk_summary", "").strip(),
                recommendations=request.form.get("recommendations", "").strip(),
                created_by=session.get("user_id"),
            )
        )
        db.session.commit()
        flash("Reporte mensual enviado a Direccion.", "success")
        return redirect(url_for("operations.reports"))

    rows = MonthlyDirectionReport.query.order_by(MonthlyDirectionReport.created_at.desc()).all()
    next_cutoff = add_business_days(date.today().replace(day=1), 4)
    return render_template("operations/reports.html", rows=rows, next_cutoff=next_cutoff)


@operations_bp.route("/acuse", methods=["GET", "POST"])
@login_required
@roles_required("ui", "contabilidad", "direccion")
def acknowledgements():
    """Registro de firma/acuse por area (UI y Contabilidad)."""
    if request.method == "POST":
        area = request.form.get("area", "").strip()
        signer = request.form.get("signer_name", "").strip()
        notes = request.form.get("notes", "").strip() or None
        if area not in {"ui", "contabilidad"} or not signer:
            flash("Datos de firma invalidos.", "error")
            return redirect(url_for("operations.acknowledgements"))
        ack = SignatureAck.query.filter_by(area=area).first()
        if not ack:
            ack = SignatureAck(area=area, signer_name=signer, notes=notes)
        else:
            ack.signer_name = signer
            ack.notes = notes
            ack.signed_at = datetime.now(timezone.utc)
        db.session.add(ack)
        db.session.commit()
        flash("Acuse firmado/actualizado.", "success")
        return redirect(url_for("operations.acknowledgements"))

    rows = SignatureAck.query.order_by(SignatureAck.signed_at.desc()).all()
    return render_template("operations/ack.html", rows=rows)
